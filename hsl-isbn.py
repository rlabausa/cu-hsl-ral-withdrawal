from fileinput import filename
import os
from time import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string       

from isbncommon import *

def main():
    # Useful constants for ws1 (original worksheet)
    WS1_KEEP_REMOVE = column_index_from_string('A')
    WS1_RELATED_PRINT_DISPLAY_CALL_NUMBER = column_index_from_string('D')
    WS1_RELATED_PRINT_TITLE = column_index_from_string('F')
    WS1_RELATED_PRINT_BARCODE = column_index_from_string('BE')
    WS1_RELATED_PRINT_PUBLICATION_YEAR = column_index_from_string('O')
    WS1_ISBN = column_index_from_string('BD')

    # Useful constants for ws2 (reviewed worksheet)
    # WS2_KEEP_REMOVE = column_index_from_string('A')
    WS2_RELATED_PRINT_DISPLAY_CALL_NUMBER = column_index_from_string('A')
    WS2_RELATED_PRINT_TITLE = column_index_from_string('B')
    WS2_RELATED_PRINT_PUBLICATION_YEAR = column_index_from_string('C')
    WS2_RELATED_PRINT_BARCODE = 'G'
    WS2_ISBN = 'H'

    # Load the workbooks
    wb1 = load_workbook('input/hsl_original.xlsx')
    wb2 = load_workbook('input/hsl_final.xlsx')

    # Target the desired worksheets
    ws1 = wb1['Print Books to Review']
    ws2 = wb2['EbookDups&ISBN - Withdraw ']

    start = time()

    for row in ws2.iter_rows(min_row = 2):
        row_num = str(row[0].row)

        title = row[WS2_RELATED_PRINT_DISPLAY_CALL_NUMBER - 1].value
        result_row = find_row_from_title(title, ws1, WS1_RELATED_PRINT_DISPLAY_CALL_NUMBER)

        if(result_row is not None):
            barcode = result_row[WS1_RELATED_PRINT_BARCODE - 1].value
            isbn = result_row[WS1_ISBN - 1].value
            ws2[WS2_RELATED_PRINT_BARCODE + row_num] = barcode
            ws2[WS2_ISBN + row_num] = isbn

        # Show progress as the file is processing
        print(row_num)

    end = time()
    
     # Save workbook
    current_datetime = datetime.now().strftime('%m-%d-%Y %H_%M_%S')
    filename = f'output/HSL REMOVE UPDATED {current_datetime}.xlsx'

    wb2.save(filename = filename)

    print(f'File saved to {Path(os.getcwd(), filename)}')
    print(f'[Processing time: {end - start} sec]')

def count_blanks():
    WS2_RELATED_PRINT_DISPLAY_CALL_NUMBER = column_index_from_string('A')
    WS2_RELATED_PRINT_TITLE = column_index_from_string('B')
    WS2_RELATED_PRINT_PUBLICATION_YEAR = column_index_from_string('C')
    WS2_RELATED_PRINT_BARCODE = column_index_from_string('G')
    WS2_ISBN = column_index_from_string('H')
    wb2 = load_workbook('output/HSL WITHDRAWAL UPDATED.xlsx')
    ws2 = wb2['EbookDups&ISBN - Withdraw ']

    counter = 0

    for row in ws2.iter_rows(min_row = 2):
        row_num = str(row[0].row)
        isbn = row[WS2_ISBN - 1].value
        if(isbn is None or isbn == ''):
            counter += 1

        # Show progress as the file is processing
        print(row_num)
    
    print(f'NUMBER BLANK: {counter}')




if __name__ == '__main__':
    main()







