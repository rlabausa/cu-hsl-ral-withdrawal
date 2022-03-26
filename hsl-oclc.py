from fileinput import filename
import os
from time import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string       

from isbncommon import *

def main():
    # Useful constants for ws1 (original worksheet containing all columns)
    WS1_RELATED_PRINT_BARCODE = column_index_from_string('Y')
    WS1_RELATED_PRINT_OCLC_NUMBER = column_index_from_string('U')

    # Useful constants for ws2 (reviewed worksheet containing subset)
    # WS2_KEEP_REMOVE = column_index_from_string('A')
    WS2_BARCODE = column_index_from_string('A')
    WS2_OCLC_NUMBER = 'B'

    dir_path = Path(os.getcwd(),'BarcodesScannedLists')
    files = get_xlsx_files(dir_path)
    print(files)

    
    for f in files:
        print(f.stem)

        # Load the workbooks
        wb1 = load_workbook('input-oclc/Books_withEbooks_REMOVE_HSL.xlsx')
        wb2 = load_workbook(Path(dir_path, f.name))

        # Target the desired worksheets
        ws1 = wb1['Ebook Duplicates']
        ws2 = wb2['Sheet1']

        start = time()

        for row in ws2.iter_rows(min_row = 2):
            row_num = str(row[0].row)

            barcode = str(row[WS2_BARCODE - 1].value) 
            oclc = find_desired_val_from_search_val(barcode, ws1, WS1_RELATED_PRINT_BARCODE, WS1_RELATED_PRINT_OCLC_NUMBER)

            if(
            oclc is not None and 
            oclc != ''):
                ws2[WS2_OCLC_NUMBER + row_num] = int(oclc)

            # Show progress as the file is processing
            print(row_num)

        end = time()
        
        # Save workbook
        current_datetime = datetime.now().strftime('%m-%d-%Y %H_%M_%S')
        output_file = f'output-oclc/{f.stem} UPDATED {current_datetime}.xlsx'

        wb2.save(filename = output_file)

        print(f'File saved to {Path(os.getcwd(), output_file)}')
        print(f'[Processing time: {end - start} sec]')



if __name__ == '__main__':
    main()
