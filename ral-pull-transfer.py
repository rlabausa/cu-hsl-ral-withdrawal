from fileinput import filename
import os
from time import time
from datetime import datetime
from pathlib import Path
import sys

from openpyxl import load_workbook
import openpyxl
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from isbncommon import *

lib_guide_dir = Path(os.getcwd(), "input", "lib_guide_title_lists")
faculty_keep_dir = Path(os.getcwd(), "input", "faculty_keep_lists", "faculty_keep_masterlist.xlsx")
pull_withdraw_list_dir = Path(os.getcwd(), "output", "pull_withdraw_lists")

# master list of all faculty requests compiled via email
faculty_keep_wb = load_workbook(Path(faculty_keep_dir))
faculty_keep_ws = faculty_keep_wb.worksheets[0]


def main():
    CALLNUMBER_COL_FACULTY = None
    BARCODE_COL_FACULTY = None
    ISBN_COL_FACULTY = None
    TRANSFERNAME_COL_FACULTY = None
    KEEP_IN_COLLECTION_FACULTY = None

    # master list of all faculty requests found in the withdrawal lists that will be transfered to their offices
    pull_transfer_wb = openpyxl.Workbook()
    pull_transfer_ws = pull_transfer_wb.active

    # set up header row into "pull & transfer" master list
    pull_transfer_ws.append(
        [
            "Keep in Collection? (Yes/No)",
            "Display Call Number",
            "Title",
            "ISBN",
            "Barcode",
            "Worldcat OCLC Number",
            "Faculty",
        ]
    )

    # get the columns
    for row in faculty_keep_ws.iter_rows(max_row=1):
        for cell in row:
            col_header = cell.value
            if col_header is not None:
                col_header = col_header.lower().strip()
                if col_header == "faculty":
                    TRANSFERNAME_COL_FACULTY = cell.column
                elif col_header == "barcode":
                    BARCODE_COL_FACULTY = cell.column
                elif (
                    col_header == "keep in collection (Yes/No)" or
                    col_header == "keep in collection? (yes/no)"
                    or col_header == "keep in collection?"
                    or col_header == "keep in collection"
                ):
                    KEEP_IN_COLLECTION_FACULTY = cell.column

    # stop execution if required columns are not present
    if TRANSFERNAME_COL_FACULTY is None or TRANSFERNAME_COL_FACULTY == "":
        print("Error: Keep masterlist file is missing Faculty column.")
        return
    elif BARCODE_COL_FACULTY is None or BARCODE_COL_FACULTY == "":
        print("Error: Keep masterlist is missing Barcode column.")
        return
    elif KEEP_IN_COLLECTION_FACULTY is None or KEEP_IN_COLLECTION_FACULTY == "":
        print("Error: Keep masterlist is missing Keep in Collection? column.")
        return

    # iterate through all categorized withdrawal files to create final "pull & withdraw" lists
    lib_guide_files = get_xlsx_files(lib_guide_dir)

    for f in lib_guide_files:
        print(f.name)

        # original version
        lib_guide_wb = load_workbook(Path(lib_guide_dir, f.name))
        lib_guide_ws = lib_guide_wb.worksheets[0]

        # final version (with faculty requests removed)
        pull_withdraw_wb = openpyxl.Workbook()
        pull_withdraw_ws = pull_withdraw_wb.active

        CALLNUMBER_COL_LIBGUIDE = None
        TITLE_COL_LIBGUIDE = None
        ISBN_COL_LIBGUIDE = None
        BARCODE_COL_LIBGUIDE = None
        WORLDCAT_COL_LIBGUIDE = None

        for row in lib_guide_ws.iter_rows(max_row=1):
            # copy header row into updated withdrawal lists
            row_vals = [cell.value for cell in row]
            pull_withdraw_ws.append(row_vals)

            # locate the necessary columns they are not in the same location across spreadsheets
            for cell in row:
                col_header = cell.value
                if col_header is not None:
                    col_header = col_header.lower().strip()
                    if col_header == "display call number":
                        CALLNUMBER_COL_LIBGUIDE = cell.column
                    elif col_header == "title":
                        TITLE_COL_LIBGUIDE = cell.column
                    if col_header == "barcode":
                        BARCODE_COL_LIBGUIDE = cell.column
                    elif col_header == "isbn":
                        ISBN_COL_LIBGUIDE = cell.column
                    elif col_header == "worldcat oclc number":
                        WORLDCAT_COL_LIBGUIDE = cell.column

        # move onto next file if there is no barcode data
        if BARCODE_COL_LIBGUIDE is None:
            continue

        # find the barcode in the faculty list
        for row in lib_guide_ws.iter_rows(min_row=2):
            barcode = row[BARCODE_COL_LIBGUIDE - 1].value
            result = find_row_from_search_val(
                barcode, faculty_keep_ws, BARCODE_COL_FACULTY
            )

            # faculty has requested book, so add it to valid "pull & transfer" list
            if result is not None:
                print(f"[FOUND] requester: {result[TRANSFERNAME_COL_FACULTY - 1].value} file: {f.name} barcode: {barcode}")

                should_keep_in_collection = result[KEEP_IN_COLLECTION_FACULTY - 1].value

                if (
                    should_keep_in_collection is None
                    or (should_keep_in_collection is not None and should_keep_in_collection.lower().strip() == "no")
                ):
                    # format specifically for the pull/transfer file
                    row_vals = [
                        should_keep_in_collection,
                        row[CALLNUMBER_COL_LIBGUIDE - 1].value,
                        row[TITLE_COL_LIBGUIDE - 1].value,
                        row[ISBN_COL_LIBGUIDE - 1].value,
                        row[BARCODE_COL_LIBGUIDE - 1].value,
                        row[WORLDCAT_COL_LIBGUIDE - 1].value,
                        result[BARCODE_COL_FACULTY - 1].value,
                    ]

                    # add row values to the "pull & transfer" spreadsheet
                    pull_transfer_ws.append(row_vals)

            # faculty has not requested book (OR faculty did not format request properly), so add it to "pull & withdraw" list
            else:
                row_vals = [cell.value for cell in row]

                # add row to the "pull & withdraw" spreadsheet
                pull_withdraw_ws.append(row_vals)

        # save every final "pull & withdraw" list
        pull_withdraw_wb.save(filename=os.path.join(pull_withdraw_list_dir, f.name))

        lib_guide_wb.close()

    # save final "pull & transfer" list
    pull_transfer_wb.save(
        filename=os.path.join(
            os.getcwd(), "output", "pull_transfer_lists", "pull_transfer.xlsx"
        )
    )
    pull_transfer_wb.close()


def log_not_found():
    BARCODE_COL_FACULTY = None
    TRANSFERNAME_COL_FACULTY = None
    KEEP_IN_COLLECTION_FACULTY = None
    BARCODE_COL_TRANSFER = None

    pull_transfer_file = Path("output", "pull_transfer_lists", "pull_transfer.xlsx")
    pull_transfer_wb = load_workbook(pull_transfer_file)
    pull_transfer_ws = pull_transfer_wb.worksheets[0]

    not_found_log_wb = openpyxl.Workbook()
    not_found_log_ws = not_found_log_wb.active

    # copy header row into log
    for row in faculty_keep_ws.iter_rows(max_row=1):
        row_vals = [cell.value for cell in row]
        not_found_log_ws.append(row_vals)

    # get the columns
    for row in faculty_keep_ws.iter_rows(max_row=1):
        for cell in row:
            col_header = cell.value
            if col_header is not None:
                col_header = col_header.lower().strip()
                if col_header == "faculty":
                    TRANSFERNAME_COL_FACULTY = cell.column
                elif col_header == "barcode":
                    BARCODE_COL_FACULTY = cell.column
                elif (
                col_header == "keep in collection (yes/no)" or 
                col_header == "keep in collection? (yes/no)" or 
                col_header == "keep in collection?" or
                col_header == "keep in collection"
                ):
                    KEEP_IN_COLLECTION_FACULTY = cell.column

    for row in pull_transfer_ws.iter_rows(max_row=1):
        for cell in row:
            col_header = cell.value
            if col_header is not None:
                col_header = col_header.lower()
                if col_header == "barcode":
                    BARCODE_COL_TRANSFER = cell.column

    for row in faculty_keep_ws.iter_rows(min_row=2):

        # we can only search if it is supposed to be transferred to their office
        should_keep_in_collection = row[KEEP_IN_COLLECTION_FACULTY - 1].value

        if(should_keep_in_collection is None or should_keep_in_collection.lower().strip() == "no"):
            barcode = row[BARCODE_COL_FACULTY - 1].value
            result = find_desired_val_from_search_val(
                barcode, pull_transfer_ws, BARCODE_COL_TRANSFER, BARCODE_COL_TRANSFER
            )

            if result is None:
                print(
                    "[NOT FOUND]",
                    row[0].row,
                    row[BARCODE_COL_FACULTY - 1].value,
                    row[TRANSFERNAME_COL_FACULTY - 1].value
                )
                row_vals = [cell.value for cell in row]
                not_found_log_ws.append(row_vals)

    current_datetime = datetime.now().strftime("%m-%d-%Y %H_%M_%S")
    not_found_log_wb.save(
        filename=os.path.join(
            "output", "log_lists", f"not_found_log {current_datetime}.xlsx"
        )
    )
    not_found_log_wb.close()
    pull_transfer_wb.close()


if __name__ == "__main__":
    start = time()
    if len(sys.argv) > 1:
        if sys.argv[1] == "--include-log" or sys.argv[1] == "-l":
            main()
            log_not_found()
        elif sys.argv[1] == "--log-only" or sys.argv[1] == "-lo":
            log_not_found()
        else:
            print("ERROR: Command not recognized.")
    else:
        main()
    end = time()
    print(f"[Processing time: {end - start} sec]")
