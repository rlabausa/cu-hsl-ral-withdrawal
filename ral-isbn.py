from fileinput import filename
import os
from time import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string       

from isbncommon import *

def main():
    # Useful constants for ws1 (original worksheet)
    WS1_KEEP_REMOVE = column_index_from_string('A')
    WS1_DISPLAY_CALL_NUMBER = column_index_from_string('E')
    WS1_BARCODE = column_index_from_string('BF')
    WS1_ISBN = column_index_from_string('BE')
    WS1_PUBLICATION_YEAR = column_index_from_string('P')
    WS1_TITLE = column_index_from_string('G')

    # Useful constants for ws2 (reviewed worksheet)
    WS2_KEEP_REMOVE = column_index_from_string('A')
    WS2_RELATED_PRINT_DISPLAY_CALL_NUMBER = column_index_from_string('D')
    WS2_RELATED_PRINT_TITLE = column_index_from_string('G')
    WS2_RELATED_PRINT_BARCODE = column_index_from_string('Y')
    WS2_RELATED_PRINT_PUBLICATION_YEAR = column_index_from_string('K')
    WS2_ISBN = 'AB'

    # Load the workbooks
    wb1 = load_workbook('input/ral_original.xlsx')
    wb2 = load_workbook('input/ral_reviewed.xlsx')

    # Target the desired worksheets
    ws1 = wb1['Print Books']
    ws2 = wb2['Ebook duplicates']

    start = time()

    for row in ws2.iter_rows(min_row = 2):
        rowNum = str(row[0].row)
        barcode = row[WS2_RELATED_PRINT_BARCODE - 1].value
        isbn = find_isbn_from_barcode(barcode, ws1, WS1_BARCODE, WS1_ISBN)
        ws2[WS2_ISBN + rowNum] = isbn

        # Show progress as the file is processing
        print(rowNum)

    end = time()
    
     # Save workbook
    current_datetime = datetime.now().strftime('%m-%d-%Y %H_%M_%S')
    filename = f'output/RAL {current_datetime}.xlsx'

    wb2.save(filename = filename)

    print(f'File saved to {Path(os.getcwd(), filename)}')
    print(f'[Processing time: {end - start} sec]')

main()








