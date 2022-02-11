import os
from time import perf_counter
from datetime import datetime
from pathlib import Path
import multiprocessing as mp

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string   

# Useful constants for ws1 (original worksheet)
WS1_KEEP_REMOVE = column_index_from_string('A')
WS1_DISPLAY_CALL_NUMBER = column_index_from_string('E')
WS1_BARCODE = column_index_from_string('BF')
WS1_ISBN = column_index_from_string('BE')
WS1_PUBLICATION_YEAR = column_index_from_string('P')
WS1_TITLE = column_index_from_string('G')

# Useful constants for ws2 (reviewed worksheet)
WS2_KEEP_REMOVE = column_index_from_string('A')
WS2_RELATED_PRINT_DISPLAY_CALL_NUMBER = column_index_from_string('D')
WS2_RELATED_PRINT_TITLE = column_index_from_string('G')
WS2_RELATED_PRINT_BARCODE = column_index_from_string('Y')
WS2_RELATED_PRINT_PUBLICATION_YEAR = column_index_from_string('K')
WS2_ISBN = column_index_from_string('AB')

# Load the workbooks
wb1 = load_workbook(Path('input', 'ral_original.xlsx'))
wb2 = load_workbook(Path('input', 'ral_reviewed.xlsx'))

# Target the desired worksheets
ws1 = wb1['Print Books']
ws2 = wb2['Ebook duplicates']

class ws_cell(object):
    def __init__(self, row, column, value):
        self.row = row
        self.column = column
        self.value = value

def find_isbn_from_row(row_with_barcode):
    """Returns the ISBN from the original worksheet if found, None otherwise.

    Keyword arguments:
    row_with_barcode -- openpyxl tuple of Cells from a worksheet
    """    
    barcode_cell = row_with_barcode[WS2_RELATED_PRINT_BARCODE - 1]
    barcode_to_find = barcode_cell.value
    for row in ws1.rows:
        barcode = row[WS1_BARCODE - 1].value

        if(barcode == barcode_to_find):
            isbn = row[WS1_ISBN - 1].value
            # print(barcode_cell.row, isbn, os.getpid())
            c = ws_cell(barcode_cell.row, WS2_ISBN, isbn)
            return c
    return None

def main():
    search_start = perf_counter()
    rows = list(ws2.rows)
    with mp.Pool(os.cpu_count()) as pool:
        results = pool.map(find_isbn_from_row, rows)
    search_end = perf_counter()
    
    save_start = perf_counter()
    for res in results:
        if(res is not None):
            print(res.row, res.column, res.value)
            ws2.cell(res.row, res.column).value = res.value    

    current_datetime = datetime.now().strftime('%m-%d-%Y %H_%M_%S')
    
    filename = Path('output', f'RAL {current_datetime}.xlsx')
    wb2.save(filename = filename)

    save_end = perf_counter()

    print(f'Search completed in {search_end - search_start} second(s)')
    print(f'Save completed in {save_end - save_start} second(s)')
        
if(__name__ == '__main__'):
    main()



   